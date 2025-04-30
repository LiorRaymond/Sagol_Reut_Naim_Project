import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

OUTPUT_DIR = "colored_results"

def color_csv(input_csv_path, output_name, matrix_type):
    # Load the correlation matrix from CSV
    df = pd.read_csv(input_csv_path, index_col=0)

    df = df.sort_index(axis=0).sort_index(axis=1)

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # ws.title = "Correlation Matrix"

    # Write headers
    headers = [''] + list(df.columns)  # Empty cell for top-left corner
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # Write data with row indices
    for row_idx, (idx_name, row_data) in enumerate(df.iterrows(), start=2):
        # Write row index name
        ws.cell(row=row_idx, column=1).value = idx_name
        
        # Write row data
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Apply conditional formatting - skip coloring if value is exactly 1 (self-correlation)
            if matrix_type == "r":
                highlight = abs(value) > 0.9
            elif matrix_type == "p":
                highlight = abs(value) < 0.1
            else:
                raise ValueError("Invalid matrix type. Use 'r' for correlation or 'p' for p-value.")

            if highlight and (row_idx != col_idx):
                red_hex = f"{int(255 * value):02X}"
                cell.fill = PatternFill(start_color=f"{red_hex}AAAA", end_color=f"{red_hex}AAAA", fill_type="solid")

    ws.freeze_panes = 'B2' 

    # Save to Excel in the main folder
    output_path = os.path.join(OUTPUT_DIR, output_name)
    wb.save(output_path)

    print(f"✅ Excel file saved successfully as: {output_path}")

color_csv('results/correlation_matrix.csv', 'r_correlation_colored.xlsx', 'r')
color_csv('results/p_fdr_feature_feature_matrix.csv', 'p_correlation_colored.xlsx', 'p')
color_csv('without_outliers/results/correlation_matrix_without_outliers.csv', 'r_correlation_colored_without_outliers.xlsx', 'r')
color_csv('without_outliers/results/p_fdr_feature_feature_matrix_without_outliers.csv', 'p_correlation_colored_without_outliers.xlsx', 'p')